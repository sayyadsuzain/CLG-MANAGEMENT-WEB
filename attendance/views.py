import logging

logger = logging.getLogger('attendance')
logger.info('Test log entry - verifying logging configuration')
logger.info('Attendance app views module loaded successfully')

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from datetime import date, datetime, timedelta
from .models import User, Student, Faculty, Course, AttendanceSession, AttendanceRecord, Assignment, Grade, Notice, AssignmentSubmission, Resource, Notification
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.core.mail import send_mail
import csv
from django.conf import settings
import os
from .forms import LoginForm, PasswordChangeForm, PasswordResetForm, SetPasswordForm
from itertools import chain
from operator import attrgetter
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.template.loader import render_to_string
from django.urls import reverse
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.views import PasswordResetView, PasswordResetDoneView, PasswordResetConfirmView, PasswordResetCompleteView
from django.utils.crypto import get_random_string
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import logging
from django.views.decorators.http import require_http_methods

# Get logger for the attendance app
logger = logging.getLogger('attendance')

def index(request):
    try:
        logger.info('Accessing index page')
        context = {}
        if request.user.is_authenticated:
            logger.debug(f'Authenticated user: {request.user.email}')
            if request.user.user_type == 'student':
                student = Student.objects.filter(user=request.user).first()
                if student:
                    context['student'] = student
                    logger.info(f'Student {student.student_id} accessed index page')
            elif request.user.user_type == 'faculty':
                faculty = Faculty.objects.filter(user=request.user).first()
                if faculty:
                    context['faculty'] = faculty
                    logger.info(f'Faculty {faculty.faculty_id} accessed index page')
        return render(request, 'index.html', context)
    except Exception as e:
        logger.error(f"Error in index view: {str(e)}", exc_info=True)
        messages.error(request, "An error occurred while loading the page. Please try again.")
        return render(request, 'index.html', {})

@csrf_protect
def login_view(request):
    logger.info('Accessing login page')
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            try:
                logger.debug(f'Attempting login for user: {email}')
                user = authenticate(username=email, password=password)
                if user is not None:
                    login(request, user)
                    logger.info(f'User {email} logged in successfully')
                    if user.user_type == 'faculty':
                        return redirect('attendance:faculty_dashboard')
                    elif user.user_type == 'student':
                        return redirect('attendance:student_dashboard')
                    else:
                        return redirect('attendance:admin_dashboard')
                else:
                    logger.warning(f'Failed login attempt for user: {email}')
                    messages.error(request, 'Invalid email or password')
            except Exception as e:
                logger.error(f'Login error for {email}: {str(e)}', exc_info=True)
                messages.error(request, 'An error occurred during login')
    else:
        form = LoginForm()
    return render(request, 'auth/login.html', {'form': form})

@csrf_protect
def register(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        user_type = request.POST.get('user_type')
        department = request.POST.get('department')
        
        # Validate required fields
        if not all([email, password, first_name, last_name, user_type, department]):
            messages.error(request, 'All fields are required')
            return render(request, 'auth/register.html')
        
        try:
            # Check if user already exists
            if User.objects.filter(email=email).exists():
                messages.error(request, 'Email already registered')
                return render(request, 'auth/register.html')
            
            # Create user
            user = User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                user_type=user_type
            )
            
            # Create profile based on user type
            if user_type == 'student':
                Student.objects.create(
                    user=user,
                    student_id=f"STU{user.id:04d}",
                    department=department
                )
            elif user_type == 'faculty':
                Faculty.objects.create(
                    user=user,
                    faculty_id=f"FAC{user.id:04d}",
                    department=department
                )
            
            messages.success(request, 'Registration successful. Please login.')
            return redirect('attendance:login')
            
        except Exception as e:
            messages.error(request, f'Registration failed: {str(e)}')
            return render(request, 'auth/register.html')
    
    return render(request, 'auth/register.html')

@login_required
def logout_view(request):
    logout(request)
    return redirect('attendance:login')

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.POST)
        if form.is_valid():
            user = request.user
            if user.check_password(form.cleaned_data['old_password']):
                user.set_password(form.cleaned_data['new_password'])
                user.save()
                messages.success(request, 'Password changed successfully')
                return redirect('attendance:login')
            else:
                messages.error(request, 'Invalid old password')
    else:
        form = PasswordChangeForm()
    return render(request, 'auth/change_password.html', {'form': form})

@login_required
def student_dashboard(request):
    if not request.user.user_type == 'student':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    try:
        student = Student.objects.get(user=request.user)
        courses = Course.objects.filter(students=student)
        
        # Get notices for student's courses and general notices
        notices = Notice.objects.filter(
            Q(course__in=courses) | Q(course__isnull=True)
        ).order_by('-created_at')[:5]
        
        # Calculate overall attendance
        attendance_records = AttendanceRecord.objects.filter(student=student)
        total_sessions = attendance_records.count()
        present_sessions = attendance_records.filter(status=True).count()
        attendance_percentage = (present_sessions / total_sessions * 100) if total_sessions > 0 else 0
        
        # Get pending assignments
        pending_assignments = Assignment.objects.filter(
            course__in=courses,
            due_date__gt=timezone.now()
        ).exclude(
            assignmentsubmission__student=student
        ).count()
        
        # Get recent submissions
        recent_submissions = AssignmentSubmission.objects.filter(
            student=student
        ).select_related('assignment').order_by('-submitted_at')[:5]
        
        # Get grades for all courses
        course_grades = []
        for course in courses:
            submissions = AssignmentSubmission.objects.filter(
                student=student,
                assignment__course=course,
                marks__isnull=False
            )
            if submissions.exists():
                avg_grade = sum(sub.marks for sub in submissions) / submissions.count()
                course_grades.append({
                    'course': course,
                    'grade': round(avg_grade, 2)
                })
        
        context = {
            'student': student,
            'courses': courses,
            'total_courses': courses.count(),
            'notices': notices,
            'attendance_percentage': round(attendance_percentage, 1),
            'pending_assignments': pending_assignments,
            'recent_submissions': recent_submissions,
            'course_grades': course_grades,
            'total_sessions': total_sessions,
            'present_sessions': present_sessions,
            'absent_sessions': total_sessions - present_sessions
        }
        return render(request, 'student/dashboard.html', context)
    except Student.DoesNotExist:
        messages.error(request, 'Student profile not found')
        return redirect('attendance:login')

@login_required
def faculty_dashboard(request):
    if not request.user.user_type == 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    try:
        faculty = Faculty.objects.get(user=request.user)
        courses = Course.objects.filter(faculty=faculty)
        
        # Calculate statistics
        total_courses = courses.count()
        total_students = Student.objects.filter(courses__in=courses).distinct().count()
        active_sessions = AttendanceSession.objects.filter(
            course__in=courses,
            date=timezone.now().date()
        ).count()
        
        # Get pending tasks (ungraded assignments)
        pending_tasks = AssignmentSubmission.objects.filter(
            assignment__course__in=courses,
            marks__isnull=True
        ).count()
        
        # Get notices for faculty's courses and general notices
        notices = Notice.objects.filter(
            Q(course__in=courses) | Q(course__isnull=True)
        ).order_by('-created_at')[:5]
        
        # Get recent activities
        attendance_records = AttendanceRecord.objects.filter(
            session__course__in=courses
        ).select_related('student__user', 'session__course').order_by('-marked_at')[:10]
        
        assignment_submissions = AssignmentSubmission.objects.filter(
            assignment__course__in=courses
        ).select_related('student__user', 'assignment').order_by('-submitted_at')[:10]
        
        # Combine and sort activities using a custom key function
        def get_activity_date(activity):
            if hasattr(activity, 'marked_at'):
                return activity.marked_at
            return activity.submitted_at
        
        recent_activities = sorted(
            chain(attendance_records, assignment_submissions),
            key=get_activity_date,
            reverse=True
        )[:10]
        
        context = {
            'faculty': faculty,
            'courses': courses,
            'total_courses': total_courses,
            'total_students': total_students,
            'active_sessions': active_sessions,
            'pending_tasks': pending_tasks,
            'notices': notices,
            'recent_activities': recent_activities,
        }
        
        return render(request, 'faculty/dashboard.html', context)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty profile not found')
        return redirect('attendance:login')

@login_required
def view_all_students(request):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    faculty = request.user.faculty
    courses = Course.objects.filter(faculty=faculty)
    
    # Get all unique students from faculty's courses
    students = Student.objects.filter(courses__in=courses).distinct()
    
    # Calculate statistics for each student
    student_stats = {}
    for student in students:
        # Calculate attendance percentage
        attendance_records = AttendanceRecord.objects.filter(
            student=student,
            session__course__in=courses
        )
        total_sessions = attendance_records.count()
        present_sessions = attendance_records.filter(status=True).count()
        attendance_percentage = (present_sessions / total_sessions * 100) if total_sessions > 0 else 0
        
        # Get enrolled courses
        enrolled_courses = courses.filter(students=student)
        
        student_stats[student.id] = {
            'attendance': round(attendance_percentage, 1),
            'courses': enrolled_courses,
            'pending_assignments': AssignmentSubmission.objects.filter(
                student=student,
                assignment__course__in=courses,
                marks__isnull=True
            ).count()
        }
    
    context = {
        'students': students,
        'student_stats': student_stats,
        'faculty': faculty
    }
    return render(request, 'faculty/view_all_students.html', context)

@login_required
def create_session(request):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    if request.method == 'POST':
        course_id = request.POST.get('course')
        date_str = request.POST.get('date')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        
        try:
            course = Course.objects.get(id=course_id, faculty=request.user.faculty)
            
            # Create the session
            session = AttendanceSession.objects.create(
                course=course,
                date=date_str,
                start_time=start_time,
                end_time=end_time,
                created_by=request.user.faculty
            )
            
            # If it's an AJAX request, return JSON response
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                students = []
                for student in course.students.all():
                    students.append({
                        'id': student.id,
                        'name': f"{student.user.first_name} {student.user.last_name}",
                        'student_id': student.student_id,
                        'email': student.user.email
                    })
                
                return JsonResponse({
                    'success': True,
                    'session_id': session.id,
                    'students': students
                })
            
            # Otherwise, redirect to mark attendance page
            return redirect('attendance:mark_attendance', session_id=session.id)
            
        except Course.DoesNotExist:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': 'Course not found'}, status=404)
            messages.error(request, 'Course not found')
            return redirect('attendance:faculty_dashboard')
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': str(e)}, status=500)
            messages.error(request, f'Error creating session: {str(e)}')
            return redirect('attendance:faculty_dashboard')
    
    faculty = request.user.faculty
    courses = Course.objects.filter(faculty=faculty)
    
    context = {
        'courses': courses,
        'today': date.today()
    }
    return render(request, 'faculty/create_session.html', context)

@login_required
def mark_attendance(request, session_id=None, course_id=None):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    # If course_id is provided, create a new session for today
    if course_id:
        faculty = request.user.faculty
        try:
            course = Course.objects.get(id=course_id, faculty=faculty)
            
            # Create a new session for today
            session = AttendanceSession.objects.create(
                course=course,
                date=date.today(),
                start_time=timezone.now().time(),
                end_time=timezone.now().replace(hour=23, minute=59, second=59).time(),
                created_by=faculty
            )
            
            return redirect('attendance:mark_attendance', session_id=session.id)
            
        except Course.DoesNotExist:
            messages.error(request, 'Course not found')
            return redirect('attendance:faculty_dashboard')
    
    # If session_id is provided, show the attendance form for that session
    if session_id:
        try:
            session = AttendanceSession.objects.get(id=session_id)
            course = session.course
            
            # Check if the faculty is assigned to this course
            if course.faculty != request.user.faculty:
                messages.error(request, 'You are not authorized to mark attendance for this session')
                return redirect('attendance:faculty_dashboard')
            
            # Get all students enrolled in the course
            students = course.students.all()
            
            # Get existing attendance records for this session
            attendance_records = {}
            for record in AttendanceRecord.objects.filter(session=session):
                attendance_records[record.student.id] = record.status
            
            if request.method == 'POST':
                # Process the attendance form submission
                for student in students:
                    attendance_value = request.POST.get(f'student_{student.id}')
                    # Convert string value to boolean: '1' for present, '0' for absent
                    status = attendance_value == '1'
                    
                    # Update or create attendance record
                    AttendanceRecord.objects.update_or_create(
                        session=session,
                        student=student,
                        defaults={
                            'status': status,
                            'marked_by': request.user.faculty
                        }
                    )
                
                messages.success(request, 'Attendance marked successfully')
                return redirect('attendance:faculty_dashboard')
            
            context = {
                'session': session,
                'students': students,
                'attendance_records': attendance_records,
                'today': date.today()
            }
            
            return render(request, 'faculty/mark_attendance_session.html', context)
            
        except AttendanceSession.DoesNotExist:
            messages.error(request, 'Session not found')
            return redirect('attendance:faculty_dashboard')
    
    # If no session_id or course_id, show the form to create a new session
    faculty = request.user.faculty
    courses = Course.objects.filter(faculty=faculty)
    
    context = {
        'courses': courses,
        'today': date.today()
    }
    return render(request, 'faculty/mark_attendance.html', context)

@login_required
def view_attendance(request):
    if request.user.user_type != 'student':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    student = request.user.student
    course_id = request.GET.get('course')
    
    if course_id:
        course = get_object_or_404(Course, id=course_id)
        attendance_records = AttendanceRecord.objects.filter(
            student=student,
            session__course=course
        ).order_by('-session__date')
    else:
        attendance_records = AttendanceRecord.objects.filter(
            student=student
        ).order_by('-session__date')
    
    courses = Course.objects.filter(students=student)
    
    # Calculate attendance statistics
    total_sessions = attendance_records.count()
    present_count = attendance_records.filter(status=True).count()
    absent_count = total_sessions - present_count
    attendance_percentage = (present_count / total_sessions * 100) if total_sessions > 0 else 0
    
    context = {
        'courses': courses,
        'attendance_records': attendance_records,
        'selected_course_id': course_id,
        'total_sessions': total_sessions,
        'present_count': present_count,
        'absent_count': absent_count,
        'attendance_percentage': round(attendance_percentage, 1)
    }
    return render(request, 'student/view_attendance.html', context)

@login_required
def view_all_assignments(request):
    if request.user.user_type != 'student':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    student = request.user.student
    courses = Course.objects.filter(students=student)
    
    # Get all assignments from all courses the student is enrolled in
    assignments = Assignment.objects.filter(course__in=courses).order_by('-due_date')
    
    # Get all submissions by this student
    submissions = {
        submission.assignment_id: submission
        for submission in AssignmentSubmission.objects.filter(
            student=student,
            assignment__in=assignments
        )
    }
    
    context = {
        'assignments': assignments,
        'submissions': submissions,
        'courses': courses
    }
    return render(request, 'student/view_all_assignments.html', context)

@login_required
def view_assignments(request, course_id):
    if request.user.user_type != 'student':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    student = request.user.student
    course = get_object_or_404(Course, id=course_id, students=student)
    
    assignments = Assignment.objects.filter(course=course).order_by('-due_date')
    submissions = {
        submission.assignment_id: submission
        for submission in AssignmentSubmission.objects.filter(
            student=student,
            assignment__course=course
        )
    }
    
    context = {
        'course': course,
        'assignments': assignments,
        'submissions': submissions
    }
    return render(request, 'student/view_assignments.html', context)

@login_required
def view_grades(request, course_id):
    if request.user.user_type != 'student':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    student = request.user.student
    course = get_object_or_404(Course, id=course_id, students=student)
    
    submissions = AssignmentSubmission.objects.filter(
        student=student,
        assignment__course=course
    ).select_related('assignment')
    
    context = {
        'course': course,
        'submissions': submissions
    }
    return render(request, 'student/view_grades.html', context)

@login_required
def submit_assignment(request, assignment_id):
    if request.user.user_type != 'student':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    student = request.user.student
    assignment = get_object_or_404(Assignment, id=assignment_id)
    
    if request.method == 'POST':
        if 'submission_file' not in request.FILES:
            messages.error(request, 'Please select a file to upload')
            return redirect('attendance:view_course_assignments', course_id=assignment.course.id)
        
        submission_file = request.FILES['submission_file']
        
        # Validate file type
        if not submission_file.name.lower().endswith('.pdf'):
            messages.error(request, 'Only PDF files are allowed')
            return redirect('attendance:submit_assignment', assignment_id=assignment_id)
        
        # Validate file size (10MB limit)
        if submission_file.size > 10 * 1024 * 1024:  # 10MB in bytes
            messages.error(request, 'File size must be less than 10MB')
            return redirect('attendance:submit_assignment', assignment_id=assignment_id)
        
        submission, created = AssignmentSubmission.objects.get_or_create(
            assignment=assignment,
            student=student,
            defaults={'file': submission_file}
        )
        
        if not created:
            submission.file = submission_file
            submission.submitted_at = timezone.now()
            submission.save()
        
        messages.success(request, 'Assignment submitted successfully')
        return redirect('attendance:view_course_assignments', course_id=assignment.course.id)
    
    # Check if assignment is past due date
    is_past_due = timezone.now() > assignment.due_date
    
    # Get existing submission if any
    existing_submission = AssignmentSubmission.objects.filter(
        assignment=assignment,
        student=student
    ).first()
    
    context = {
        'assignment': assignment,
        'is_past_due': is_past_due,
        'existing_submission': existing_submission
    }
    return render(request, 'student/submit_assignment.html', context)

@login_required
def create_assignment(request, course_id):
    # First check if user is faculty type
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied. Only faculty members can create assignments.')
        return redirect('attendance:login')
    
    # Check if user has faculty profile
    try:
        faculty = Faculty.objects.get(user=request.user)
    except Faculty.DoesNotExist:
        messages.error(request, 'Faculty profile not found. Please contact administrator.')
        return redirect('attendance:login')
    
    try:
        course = Course.objects.get(id=course_id, faculty=faculty)
        if request.method == 'POST':
            title = request.POST.get('title')
            description = request.POST.get('description')
            due_date = request.POST.get('due_date')
            max_marks = request.POST.get('max_marks')
            assignment_file = request.FILES.get('assignment_file')
            
            # Validate required fields
            if not all([title, description, due_date, max_marks]):
                messages.error(request, 'All fields are required except assignment file.')
                return render(request, 'faculty/create_assignment.html', {'course': course})
            
            # Convert the datetime-local input to proper datetime format
            try:
                due_date = datetime.strptime(due_date, '%Y-%m-%dT%H:%M')
            except ValueError:
                messages.error(request, 'Invalid date format')
                return render(request, 'faculty/create_assignment.html', {'course': course})
            
            # Validate file type if provided
            if assignment_file and not assignment_file.name.lower().endswith('.pdf'):
                messages.error(request, 'Only PDF files are allowed')
                return render(request, 'faculty/create_assignment.html', {'course': course})
            
            try:
                assignment = Assignment.objects.create(
                    course=course,
                    title=title,
                    description=description,
                    due_date=due_date,
                    max_marks=float(max_marks),
                    created_by=faculty,
                    file=assignment_file if assignment_file else None
                )
                
                messages.success(request, 'Assignment created successfully!')
                return redirect('attendance:course_assignments', course_id=course.id)
            except ValueError:
                messages.error(request, 'Invalid marks value. Please enter a valid number.')
                return render(request, 'faculty/create_assignment.html', {'course': course})
            
        context = {
            'course': course
        }
        return render(request, 'faculty/create_assignment.html', context)
        
    except Course.DoesNotExist:
        messages.error(request, 'Course not found or you do not have permission to create assignments.')
        return redirect('attendance:faculty_dashboard')

@login_required
def grade_assignment(request, assignment_id):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    faculty = request.user.faculty
    assignment = get_object_or_404(Assignment, id=assignment_id, course__faculty=faculty)
    
    if request.method == 'POST':
        for key, value in request.POST.items():
            if key.startswith('grade_'):
                submission_id = key.split('_')[1]
                grade = value
                remarks = request.POST.get(f'remarks_{submission_id}')
                
                try:
                    submission = AssignmentSubmission.objects.get(id=submission_id)
                    if grade:
                        submission.marks = grade
                        submission.remarks = remarks
                        submission.graded_by = faculty
                        submission.graded_at = timezone.now()
                        submission.save()
                except AssignmentSubmission.DoesNotExist:
                    continue
        
        messages.success(request, 'Grades updated successfully')
        return redirect('attendance:faculty_dashboard')
    
    # Get all submissions for this assignment
    submissions = AssignmentSubmission.objects.filter(assignment=assignment)
    context = {
        'assignment': assignment,
        'submissions': submissions
    }
    return render(request, 'faculty/grade_assignment.html', context)

@login_required
def update_grades(request, course_id):
    try:
        course = Course.objects.get(id=course_id)
    except Course.DoesNotExist:
        messages.error(request, 'Course not found')
        return redirect('attendance:faculty_dashboard')
    
    if request.user.user_type != 'faculty' or course.faculty.user != request.user:
        messages.error(request, 'Access denied')
        return redirect('attendance:faculty_dashboard')
    
    if request.method == 'POST':
        # Process grade updates
        students = course.students.all()
        for student in students:
            grade_value = request.POST.get(f'grade_{student.id}')
            percentage = request.POST.get(f'percentage_{student.id}')
            remarks = request.POST.get(f'remarks_{student.id}')
            
            if grade_value and percentage:  # Only update if both grade and percentage are provided
                try:
                    percentage = float(percentage)
                    if 0 <= percentage <= 100:  # Validate percentage range
                        Grade.objects.update_or_create(
                            course=course,
                            student=student,
                            defaults={
                                'grade': grade_value,
                                'percentage': percentage,
                                'remarks': remarks or ''
                            }
                        )
                except ValueError:
                    messages.error(request, f'Invalid percentage value for student {student.user.get_full_name()}')
                    continue
        
        messages.success(request, 'Grades updated successfully')
        return redirect('attendance:view_course_students', course_id=course.id)
    
    # For GET request, show the form
    students = course.students.all()
    grade_objects = Grade.objects.filter(course=course)
    
    # Create separate dictionaries for grades, percentages, and remarks
    grades = {grade.student_id: grade.grade for grade in grade_objects}
    percentages = {grade.student_id: grade.percentage for grade in grade_objects}
    remarks = {grade.student_id: grade.remarks for grade in grade_objects}
    
    context = {
        'course': course,
        'students': students,
        'grades': grades,
        'percentages': percentages,
        'remarks': remarks
    }
    return render(request, 'faculty/update_grades.html', context)

@login_required
def create_notice(request):
    if request.user.user_type != 'faculty' and not request.user.is_superuser:
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        course_id = request.POST.get('course')
        priority = request.POST.get('priority', 'medium')
        
        # Get the faculty instance if the user is faculty
        created_by = request.user.faculty if request.user.user_type == 'faculty' else None
        
        notice = Notice.objects.create(
            title=title,
            content=content,
            course_id=course_id if course_id else None,
            created_by=created_by,
            priority=priority
        )
        
        messages.success(request, 'Notice created successfully')
        if request.user.is_superuser:
            return redirect('attendance:admin_dashboard')
        return redirect('attendance:faculty_dashboard')
    
    courses = Course.objects.all()
    if request.user.user_type == 'faculty':
        courses = courses.filter(faculty=request.user.faculty)
    
    context = {
        'courses': courses
    }
    return render(request, 'faculty/create_notice.html', context)

@login_required
def admin_dashboard(request):
    if not request.user.is_superuser:
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    context = {
        'total_students': Student.objects.count(),
        'total_faculty': Faculty.objects.count(),
        'total_courses': Course.objects.count(),
        'recent_notices': Notice.objects.order_by('-created_at')[:5]
    }
    return render(request, 'admin/dashboard.html', context)

@login_required
def get_course_students(request, course_id):
    if request.user.user_type != 'faculty':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        course = Course.objects.get(id=course_id, faculty=request.user.faculty)
        students = []
        for student in course.students.all():
            students.append({
                'id': student.id,
                'name': f"{student.user.first_name} {student.user.last_name}",
                'student_id': student.student_id,
                'email': student.user.email
            })
        
        return JsonResponse({
            'success': True,
            'students': students
        })
    except Course.DoesNotExist:
        return JsonResponse({'error': 'Course not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def add_student_to_course(request, course_id):
    if request.user.user_type != 'faculty':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=405)
    
    course = get_object_or_404(Course, id=course_id, faculty=request.user.faculty)
    student_id = request.POST.get('student_id')
    
    try:
        student = Student.objects.get(student_id=student_id)
        course.students.add(student)
        return JsonResponse({
            'success': True,
            'student': {
                'id': student.id,
                'name': f"{student.user.first_name} {student.user.last_name}",
                'student_id': student.student_id
            }
        })
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

@login_required
def remove_student_from_course(request, course_id, student_id):
    if request.user.user_type != 'faculty':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=405)
    
    try:
        course = get_object_or_404(Course, id=course_id, faculty=request.user.faculty)
        student = get_object_or_404(Student, id=student_id)
        
        if student not in course.students.all():
            return JsonResponse({'error': 'Student is not enrolled in this course'}, status=400)
        
        course.students.remove(student)
        return JsonResponse({
            'success': True,
            'message': f'Student {student.user.get_full_name()} has been removed from {course.name}'
        })
    except Course.DoesNotExist:
        return JsonResponse({'error': 'Course not found'}, status=404)
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def search_students(request):
    if request.user.user_type != 'faculty':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'students': []})
    
    students = Student.objects.filter(
        Q(user__first_name__icontains=query) |
        Q(user__last_name__icontains=query) |
        Q(student_id__icontains=query)
    )[:10]
    
    results = [{
        'id': student.id,
        'name': f"{student.user.first_name} {student.user.last_name}",
        'student_id': student.student_id
    } for student in students]
    
    return JsonResponse({'students': results})

@login_required
def student_attendance_history(request, student_id):
    if request.user.user_type != 'faculty':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    course_id = request.GET.get('course_id')
    if not course_id:
        return JsonResponse({'error': 'Course ID is required'}, status=400)
    
    try:
        student = Student.objects.get(id=student_id)
        course = Course.objects.get(id=course_id)
        
        # Get the last 10 attendance records for this student in the course
        records = AttendanceRecord.objects.filter(
            student=student,
            session__course=course
        ).order_by('-session__date')[:10]
        
        history = [{
            'date': record.session.date.strftime('%Y-%m-%d'),
            'status': record.status,
            'marked_by': record.marked_by.user.get_full_name() if record.marked_by else 'System',
            'marked_at': record.marked_at.strftime('%Y-%m-%d %H:%M:%S') if record.marked_at else None
        } for record in records]
        
        return JsonResponse({
            'success': True,
            'student_name': student.user.get_full_name(),
            'student_id': student.student_id,
            'course_name': course.name,
            'history': history
        })
    except (Student.DoesNotExist, Course.DoesNotExist):
        return JsonResponse({'error': 'Student or Course not found'}, status=404)

@login_required
def export_attendance(request, session_id=None):
    if request.user.user_type != 'faculty':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        centered = Alignment(horizontal="center")
        
        if session_id:
            # Export attendance for a specific session
            session = AttendanceSession.objects.get(id=session_id, course__faculty=request.user.faculty)
            
            # Set worksheet title
            ws.title = f"Attendance {session.date}"
            
            # Write headers
            headers = ['Student ID', 'Name', 'Status', 'Marked By', 'Marked At']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered
            
            # Get records
            records = AttendanceRecord.objects.filter(session=session).select_related(
                'student', 'student__user', 'marked_by', 'marked_by__user'
            )
            
            # Write data
            for row, record in enumerate(records, 2):
                marked_by = record.marked_by.user.get_full_name() if record.marked_by else 'System'
                marked_at = record.marked_at.strftime('%d-%m-%Y %I:%M %p') if record.marked_at else 'N/A'
                
                ws.cell(row=row, column=1, value=record.student.student_id)
                ws.cell(row=row, column=2, value=record.student.user.get_full_name())
                ws.cell(row=row, column=3, value='Present' if record.status else 'Absent')
                ws.cell(row=row, column=4, value=marked_by)
                ws.cell(row=row, column=5, value=marked_at)
            
            filename = f"attendance_{session.course.course_code}_{session.date}.xlsx"
        else:
            # Export all attendance records
            ws.title = "All Attendance Records"
            
            # Write headers
            headers = ['Course', 'Date', 'Student ID', 'Name', 'Status', 'Marked By', 'Marked At']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered
            
            # Get records
            records = AttendanceRecord.objects.filter(
                session__course__faculty=request.user.faculty
            ).select_related(
                'session', 'session__course',
                'student', 'student__user',
                'marked_by', 'marked_by__user'
            ).order_by('session__course__course_code', '-session__date')
            
            # Write data
            for row, record in enumerate(records, 2):
                marked_by = record.marked_by.user.get_full_name() if record.marked_by else 'System'
                marked_at = record.marked_at.strftime('%d-%m-%Y %I:%M %p') if record.marked_at else 'N/A'
                session_date = record.session.date.strftime('%d-%m-%Y')
                
                ws.cell(row=row, column=1, value=record.session.course.course_code)
                ws.cell(row=row, column=2, value=session_date)
                ws.cell(row=row, column=3, value=record.student.student_id)
                ws.cell(row=row, column=4, value=record.student.user.get_full_name())
                ws.cell(row=row, column=5, value='Present' if record.status else 'Absent')
                ws.cell(row=row, column=6, value=marked_by)
                ws.cell(row=row, column=7, value=marked_at)
            
            filename = "attendance_all_courses.xlsx"
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create the response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save the workbook to the response
        wb.save(response)
        return response
        
    except AttendanceSession.DoesNotExist:
        return JsonResponse({'error': 'Session not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def send_attendance_notifications(request, session_id):
    if request.user.user_type != 'faculty':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=405)
    
    try:
        session = AttendanceSession.objects.get(id=session_id, course__faculty=request.user.faculty)
        absent_records = AttendanceRecord.objects.filter(session=session, status=False).select_related('student', 'student__user')
        
        notification_count = 0
        for record in absent_records:
            try:
                send_mail(
                    subject=f'Absence Notification - {session.course.code}',
                    message=f'''Dear {record.student.user.get_full_name()},

This is to inform you that you were marked absent for {session.course.code} - {session.course.name} on {session.date}.

Class Details:
- Date: {session.date}
- Time: {session.start_time} - {session.end_time}
- Faculty: {session.created_by.user.get_full_name()}

Please contact your faculty if you believe this was marked in error.

Best regards,
{settings.COLLEGE_NAME}''',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[record.student.user.email],
                )
                notification_count += 1
            except Exception as e:
                print(f"Failed to send email to {record.student.user.email}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': f'Sent {notification_count} notification(s)'
        })
    except AttendanceSession.DoesNotExist:
        return JsonResponse({'error': 'Session not found'}, status=404)

@login_required
def chat_room(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    if request.user.user_type == 'student' and not course.students.filter(user=request.user).exists():
        messages.error(request, 'Access denied')
        return redirect('attendance:student_dashboard')
    elif request.user.user_type == 'faculty' and course.faculty.user != request.user:
        messages.error(request, 'Access denied')
        return redirect('attendance:faculty_dashboard')
    
    context = {
        'course': course,
        'room_name': f'course_{course_id}',
        'user_full_name': request.user.get_full_name(),
        'user_type': request.user.user_type
    }
    return render(request, 'chat/room.html', context)

@login_required
def upload_resource(request, course_id):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    course = get_object_or_404(Course, id=course_id, faculty=request.user.faculty)
    
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        file = request.FILES.get('resource_file')
        
        if not all([title, file]):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Title and file are required'})
            messages.error(request, 'Title and file are required')
            return redirect('attendance:upload_resource', course_id=course_id)
        
        # Ensure the file is a PDF
        if not file.name.lower().endswith('.pdf'):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Only PDF files are allowed'})
            messages.error(request, 'Only PDF files are allowed')
            return redirect('attendance:upload_resource', course_id=course_id)
            
        try:
            # Create media/resources directory if it doesn't exist
            os.makedirs(os.path.join(settings.MEDIA_ROOT, 'resources'), exist_ok=True)
            
            resource = Resource.objects.create(
                course=course,
                title=title,
                description=description,
                file=file,
                uploaded_by=request.user
            )
            
            # Send notification to all students in the course
            students = course.students.all()
            for student in students:
                Notification.objects.create(
                    user=student.user,
                    title=f'New Resource: {title}',
                    message=f'A new resource has been uploaded in {course.name}',
                    notification_type='resource',
                    related_id=resource.id
                )
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'redirect_url': reverse('attendance:course_resources', args=[course_id])
                })
            
            messages.success(request, 'Resource uploaded successfully')
            return redirect('attendance:course_resources', course_id=course_id)
            
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(request, f'Error uploading resource: {str(e)}')
            return redirect('attendance:upload_resource', course_id=course_id)
    
    context = {
        'course': course
    }
    return render(request, 'faculty/upload_resource.html', context)

@login_required
def course_resources(request, course_id):
    if request.user.user_type == 'student':
        course = get_object_or_404(Course, id=course_id, students=request.user.student)
    else:
        course = get_object_or_404(Course, id=course_id, faculty=request.user.faculty)
    
    resources = Resource.objects.filter(course=course).order_by('-uploaded_at')
    context = {
        'course': course,
        'resources': resources
    }
    return render(request, 'course/resources.html', context)

@login_required
def delete_resource(request, resource_id):
    if request.user.user_type != 'faculty':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=405)
    
    resource = get_object_or_404(Resource, id=resource_id, uploaded_by=request.user)
    resource.delete()
    
    return JsonResponse({'success': True})

@login_required
def notifications(request):
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')[:50]
    unread_count = notifications.filter(read=False).count()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'notifications': [{
                'id': n.id,
                'title': n.title,
                'message': n.message,
                'created_at': n.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'read': n.read
            } for n in notifications],
            'unread_count': unread_count
        })
    
    context = {
        'notifications': notifications,
        'unread_count': unread_count
    }
    return render(request, 'notifications/list.html', context)

@login_required
def mark_notification_read(request, notification_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=405)
    
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.read = True
    notification.save()
    
    return JsonResponse({'success': True})

@login_required
def generate_report(request):
    if not request.user.is_superuser:
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    report_type = request.GET.get('type')
    course_id = request.GET.get('course')
    
    if report_type == 'attendance':
        course = get_object_or_404(Course, id=course_id)
        records = AttendanceRecord.objects.filter(session__course=course)
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="attendance_report_{course.code}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student ID', 'Name', 'Total Sessions', 'Present', 'Absent', 'Percentage'])
        
        students = course.students.all()
        for student in students:
            total = records.filter(student=student).count()
            present = records.filter(student=student, status=True).count()
            absent = total - present
            percentage = (present / total * 100) if total > 0 else 0
            
            writer.writerow([
                student.student_id,
                student.user.get_full_name(),
                total,
                present,
                absent,
                f"{percentage:.2f}%"
            ])
        
        return response
    
    elif report_type == 'grades':
        course = get_object_or_404(Course, id=course_id)
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="grades_report_{course.code}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student ID', 'Name', 'Grade', 'Percentage', 'Remarks'])
        
        grades = Grade.objects.filter(course=course).select_related('student', 'student__user')
        for grade in grades:
            writer.writerow([
                grade.student.student_id,
                grade.student.user.get_full_name(),
                grade.grade,
                grade.percentage,
                grade.remarks
            ])
        
        return response
    
    elif report_type == 'performance':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="overall_performance_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student ID', 'Name', 'Department', 'Average Grade', 'Overall Attendance', 'Total Courses'])
        
        students = Student.objects.all()
        for student in students:
            grades = Grade.objects.filter(student=student)
            attendance_records = AttendanceRecord.objects.filter(student=student)
            
            total_percentage = sum(float(g.percentage) for g in grades if g.percentage)
            avg_grade = total_percentage / grades.count() if grades.exists() else 0
            
            total_sessions = attendance_records.count()
            present_sessions = attendance_records.filter(status=True).count()
            attendance_percentage = (present_sessions / total_sessions * 100) if total_sessions > 0 else 0
            
            writer.writerow([
                student.student_id,
                student.user.get_full_name(),
                student.department,
                f"{avg_grade:.2f}%",
                f"{attendance_percentage:.2f}%",
                student.course_set.count()
            ])
        
        return response
    
    courses = Course.objects.all()
    context = {
        'courses': courses
    }
    return render(request, 'admin/generate_report.html', context)

@login_required
def add_course(request):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    if request.method == 'POST':
        course_code = request.POST.get('course_code')
        course_name = request.POST.get('course_name')
        description = request.POST.get('description')
        
        if Course.objects.filter(course_code=course_code).exists():
            messages.error(request, 'Course code already exists')
            return redirect('attendance:faculty_dashboard')
        
        course = Course.objects.create(
            course_code=course_code,
            name=course_name,
            description=description,
            faculty=request.user.faculty
        )
        messages.success(request, 'Course added successfully')
        return redirect('attendance:faculty_dashboard')
    
    return render(request, 'faculty/add_course.html')

@login_required
def add_student(request):
    if request.user.user_type != 'faculty':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method == 'POST':
        # Get form data
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        student_id = request.POST.get('student_id')
        department = request.POST.get('department')
        semester = request.POST.get('semester')
        course_id = request.POST.get('course_id')
        
        # Validate required fields
        if not all([first_name, last_name, email, student_id, department, semester]):
            return JsonResponse({
                'error': 'All fields are required',
                'field': None
            }, status=400)
        
        try:
            # Check if email exists
            if User.objects.filter(email=email).exists():
                existing_user = User.objects.get(email=email)
                if hasattr(existing_user, 'student'):
                    # If user exists and is a student, try to add them to the course
                    try:
                        course = Course.objects.get(id=course_id, faculty=request.user.faculty)
                        if course.students.filter(id=existing_user.student.id).exists():
                            return JsonResponse({
                                'error': 'Student is already enrolled in this course',
                                'field': 'email'
                            }, status=400)
                        
                        # Add existing student to course
                        course.students.add(existing_user.student)
                        return JsonResponse({
                            'success': True,
                            'student': {
                                'id': existing_user.student.id,
                                'name': f"{existing_user.first_name} {existing_user.last_name}",
                                'student_id': existing_user.student.student_id,
                                'email': existing_user.email
                            }
                        })
                    except Course.DoesNotExist:
                        return JsonResponse({
                            'error': 'Course not found',
                            'field': 'course_id'
                        }, status=404)
                else:
                    return JsonResponse({
                        'error': 'Email already exists but is not associated with a student account',
                        'field': 'email'
                    }, status=400)
            
            # Check if student ID exists
            if Student.objects.filter(student_id=student_id).exists():
                return JsonResponse({
                    'error': 'Student ID already exists',
                    'field': 'studentId'
                }, status=400)
            
            # Create new user and student
            user = User.objects.create_user(
                username=email,
                email=email,
                password=student_id,  # Set initial password as student ID
                first_name=first_name,
                last_name=last_name,
                user_type='student'
            )
            
            # Create student profile
            student = Student.objects.create(
                user=user,
                student_id=student_id,
                department=department,
                semester=semester
            )
            
            # Add student to course if specified
            if course_id:
                try:
                    course = Course.objects.get(id=course_id, faculty=request.user.faculty)
                    course.students.add(student)
                except Course.DoesNotExist:
                    # Delete the created user and student if course doesn't exist
                    user.delete()  # This will cascade delete the student profile
                    return JsonResponse({
                        'error': 'Course not found or you do not have permission to add students to it',
                        'field': 'course_id'
                    }, status=404)
            
            return JsonResponse({
                'success': True,
                'student': {
                    'id': student.id,
                    'name': f"{student.user.first_name} {student.user.last_name}",
                    'student_id': student.student_id,
                    'email': student.user.email
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'error': str(e),
                'field': None
            }, status=500)
    
    # If not POST, return error
    return JsonResponse({'error': 'Invalid method'}, status=405)

@login_required
def manage_grades(request):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    if request.method == 'POST':
        course_id = request.POST.get('course_id')
        student_id = request.POST.get('student_id')
        grade_value = request.POST.get('grade')
        percentage = request.POST.get('percentage')
        remarks = request.POST.get('remarks', '')
        
        try:
            course = Course.objects.get(id=course_id, faculty=request.user.faculty)
            student = Student.objects.get(id=student_id)
            
            # Update or create grade
            grade, created = Grade.objects.update_or_create(
                course=course,
                student=student,
                defaults={
                    'grade': grade_value,
                    'percentage': percentage,
                    'remarks': remarks
                }
            )
            
            messages.success(request, 'Grade updated successfully')
            return redirect('attendance:course_grades', course_id=course_id)
            
        except Course.DoesNotExist:
            messages.error(request, 'Course not found')
        except Student.DoesNotExist:
            messages.error(request, 'Student not found')
        except Exception as e:
            messages.error(request, f'Error updating grade: {str(e)}')
        
        return redirect('attendance:faculty_dashboard')
    
    # GET request - show the form
    courses = Course.objects.filter(faculty=request.user.faculty)
    context = {
        'courses': courses,
        'grade_choices': Grade.GRADE_CHOICES
    }
    return render(request, 'faculty/manage_grades.html', context)

def password_reset_done(request):
    return render(request, "auth/password_reset_done.html")

def password_reset_confirm(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
        
        # Check if token is valid and not expired
        if user.password_reset_token != token:
            messages.error(request, "The password reset link is invalid.")
            return redirect("attendance:login")
            
        # Check if token is expired (24 hours)
        if user.password_reset_token_created and timezone.now() > user.password_reset_token_created + timedelta(hours=24):
            messages.error(request, "The password reset link has expired.")
            return redirect("attendance:login")
            
        if request.method == "POST":
            form = SetPasswordForm(user, request.POST)
            if form.is_valid():
                form.save()
                # Clear the reset token
                user.password_reset_token = None
                user.password_reset_token_created = None
                user.save()
                messages.success(request, "Your password has been set. You may go ahead and log in now.")
                return redirect("attendance:password_reset_complete")
        else:
            form = SetPasswordForm(user)
        return render(request, "auth/password_reset_confirm.html", {"form": form})
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        messages.error(request, "The password reset link is invalid.")
        return redirect("attendance:login")

def password_reset_complete(request):
    return render(request, "auth/password_reset_complete.html")

@login_required
def profile(request):
    if request.method == 'POST':
        # Handle profile update
        user = request.user
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        user.phone_number = request.POST.get('phone_number')
        
        if 'profile_picture' in request.FILES:
            if user.profile_picture:
                # Delete old profile picture if it exists
                try:
                    os.remove(user.profile_picture.path)
                except (FileNotFoundError, ValueError):
                    pass
            user.profile_picture = request.FILES['profile_picture']
        
        user.save()
        messages.success(request, 'Profile updated successfully')
        return redirect('attendance:profile')
    
    context = {
        'user': request.user,
        'profile_picture_url': request.user.profile_picture.url if request.user.profile_picture else None,
    }
    return render(request, 'auth/profile.html', context)

@login_required
def view_courses(request):
    if request.user.user_type != 'student':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    student = request.user.student
    courses = Course.objects.filter(students=student)
    
    context = {
        'courses': courses,
        'student': student
    }
    return render(request, 'student/view_courses.html', context)

@login_required
def course_students(request, course_id):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    course = get_object_or_404(Course, id=course_id, faculty=request.user.faculty)
    students = course.students.all()
    
    # Calculate attendance statistics for each student
    attendance_stats = {}
    for student in students:
        total_sessions = AttendanceRecord.objects.filter(
            session__course=course,
            student=student
        ).count()
        
        if total_sessions > 0:
            present_count = AttendanceRecord.objects.filter(
                session__course=course,
                student=student,
                status=True
            ).count()
            attendance_percentage = (present_count / total_sessions) * 100
        else:
            attendance_percentage = 0
            
        attendance_stats[student.id] = round(attendance_percentage, 1)
    
    context = {
        'course': course,
        'students': students,
        'attendance_stats': attendance_stats
    }
    return render(request, 'faculty/course_students.html', context)

@login_required
def password_reset_request(request):
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            try:
                user = User.objects.get(email=email)
                # Generate password reset token
                token = default_token_generator.make_token(user)
                uid = urlsafe_base64_encode(force_bytes(user.pk))
                
                # Build password reset URL
                reset_url = request.build_absolute_uri(
                    reverse('attendance:password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
                )
                
                # Send email
                send_mail(
                    'Password Reset Request',
                    f'Please click the following link to reset your password: {reset_url}',
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False
                )
                
                return redirect('attendance:password_reset_done')
            except User.DoesNotExist:
                messages.error(request, 'No user found with that email address.')
    else:
        form = PasswordResetForm()
    
    return render(request, 'auth/password_reset_form.html', {'form': form})

@login_required
def course_assignments(request, course_id):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    course = get_object_or_404(Course, id=course_id, faculty=request.user.faculty)
    assignments = Assignment.objects.filter(course=course).order_by('-due_date')
    
    # Calculate statistics for each assignment
    assignment_stats = {}
    for assignment in assignments:
        total_students = course.students.count()
        submitted_count = AssignmentSubmission.objects.filter(assignment=assignment).count()
        assignment_stats[assignment.id] = {
            'total_students': total_students,
            'submitted': submitted_count,
            'submission_percentage': round((submitted_count / total_students * 100) if total_students > 0 else 0, 1)
        }
    
    context = {
        'course': course,
        'assignments': assignments,
        'assignment_stats': assignment_stats
    }
    return render(request, 'faculty/course_assignments.html', context)

@login_required
@require_http_methods(["POST"])
@ensure_csrf_cookie
def delete_assignment(request, assignment_id):
    logger.info(f'Attempting to delete assignment {assignment_id}')
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if not is_ajax:
        logger.warning('Non-AJAX request received for delete_assignment')
        return JsonResponse({
            'error': 'Invalid request. Must be an AJAX request.'
        }, status=400)
    
    if request.user.user_type != 'faculty':
        logger.warning(f'Unauthorized delete attempt by user {request.user.email}')
        return JsonResponse({
            'error': 'Access denied. Only faculty members can delete assignments.'
        }, status=403)
    
    try:
        assignment = get_object_or_404(
            Assignment,
            id=assignment_id,
            course__faculty=request.user.faculty
        )
        
        course_id = assignment.course.id
        logger.info(f'Deleting assignment {assignment_id} from course {course_id}')
        
        assignment.delete()
        logger.info(f'Successfully deleted assignment {assignment_id}')
        
        return JsonResponse({
            'success': True,
            'message': 'Assignment deleted successfully',
            'redirect_url': reverse('attendance:course_assignments', args=[course_id])
        })
        
    except Assignment.DoesNotExist:
        logger.error(f'Assignment {assignment_id} not found or permission denied')
        return JsonResponse({
            'error': 'Assignment not found or you do not have permission to delete it.'
        }, status=404)
    except Exception as e:
        logger.error(f'Error deleting assignment {assignment_id}: {str(e)}', exc_info=True)
        return JsonResponse({
            'error': f'Failed to delete assignment: {str(e)}'
        }, status=500)

@login_required
def manage_courses(request):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    faculty = request.user.faculty
    courses = Course.objects.filter(faculty=faculty)
    
    context = {
        'courses': courses,
        'faculty': faculty
    }
    return render(request, 'faculty/manage_courses.html', context)

@login_required
def grade_submission(request, submission_id):
    if request.user.user_type != 'faculty':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=405)
    
    try:
        # Get the submission and verify faculty has access
        submission = get_object_or_404(
            AssignmentSubmission, 
            id=submission_id,
            assignment__course__faculty=request.user.faculty
        )
        
        # Parse the JSON data
        import json
        data = json.loads(request.body)
        marks = data.get('marks')
        remarks = data.get('remarks')
        
        # Validate marks
        try:
            marks = float(marks)
            if marks < 0 or marks > submission.assignment.max_marks:
                return JsonResponse({
                    'error': f'Marks must be between 0 and {submission.assignment.max_marks}'
                }, status=400)
        except (TypeError, ValueError):
            return JsonResponse({'error': 'Invalid marks value'}, status=400)
        
        # Update the submission
        submission.marks = marks
        submission.feedback = remarks
        submission.graded_by = request.user.faculty
        submission.graded_at = timezone.now()
        submission.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Grade saved successfully'
        })
        
    except AssignmentSubmission.DoesNotExist:
        return JsonResponse({'error': 'Submission not found'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def view_assignment_detail(request, assignment_id):
    if request.user.user_type != 'student':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    student = request.user.student
    assignment = get_object_or_404(Assignment, id=assignment_id)
    
    # Check if student is enrolled in the course
    if not student.courses.filter(id=assignment.course.id).exists():
        messages.error(request, 'Access denied')
        return redirect('attendance:student_dashboard')
    
    # Get student's submission for this assignment if exists
    try:
        submission = AssignmentSubmission.objects.get(
            assignment=assignment,
            student=student
        )
    except AssignmentSubmission.DoesNotExist:
        submission = None
    
    context = {
        'assignment': assignment,
        'submission': submission
    }
    return render(request, 'student/assignment_detail.html', context)

@login_required
def view_active_sessions(request):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    faculty = request.user.faculty
    today = timezone.now().date()
    
    # Get all active sessions for faculty's courses
    active_sessions = AttendanceSession.objects.filter(
        course__faculty=faculty,
        date=today
    ).select_related('course')
    
    # Get attendance statistics for each session
    session_stats = {}
    for session in active_sessions:
        total_students = session.course.students.count()
        marked_attendance = AttendanceRecord.objects.filter(session=session).count()
        present_students = AttendanceRecord.objects.filter(session=session, status=True).count()
        
        session_stats[session.id] = {
            'total_students': total_students,
            'marked_attendance': marked_attendance,
            'present_students': present_students,
            'pending_students': total_students - marked_attendance
        }
    
    context = {
        'active_sessions': active_sessions,
        'session_stats': session_stats,
        'faculty': faculty
    }
    return render(request, 'faculty/view_active_sessions.html', context)

@login_required
def view_pending_tasks(request):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    faculty = request.user.faculty
    courses = Course.objects.filter(faculty=faculty)
    
    # Get ungraded assignments
    pending_submissions = AssignmentSubmission.objects.filter(
        assignment__course__in=courses,
        marks__isnull=True
    ).select_related('student__user', 'assignment__course')
    
    # Get today's unmarked attendance
    today = timezone.now().date()
    active_sessions = AttendanceSession.objects.filter(
        course__in=courses,
        date=today
    )
    
    pending_attendance = []
    for session in active_sessions:
        total_students = session.course.students.count()
        marked_count = AttendanceRecord.objects.filter(session=session).count()
        if marked_count < total_students:
            pending_attendance.append({
                'session': session,
                'pending_count': total_students - marked_count,
                'total_students': total_students
            })
    
    context = {
        'pending_submissions': pending_submissions,
        'pending_attendance': pending_attendance,
        'faculty': faculty
    }
    return render(request, 'faculty/view_pending_tasks.html', context)

@login_required
def view_all_notices(request):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    faculty = request.user.faculty
    courses = Course.objects.filter(faculty=faculty)
    
    # Get all notices for faculty's courses and general notices
    notices = Notice.objects.filter(
        Q(course__in=courses) | Q(course__isnull=True)
    ).order_by('-created_at')
    
    context = {
        'notices': notices,
        'faculty': faculty,
        'courses': courses
    }
    return render(request, 'faculty/view_all_notices.html', context)

@login_required
def delete_notice(request, notice_id):
    if request.user.user_type != 'faculty':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=405)
    
    try:
        notice = Notice.objects.filter(
            Q(course__faculty=request.user.faculty) | Q(created_by=request.user.faculty),
            id=notice_id
        ).first()
        
        if notice:
            notice.delete()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'error': 'Notice not found or access denied'}, status=404)
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def student_details(request, student_id):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    faculty = request.user.faculty
    try:
        # Get student and ensure they are in one of faculty's courses
        student = Student.objects.select_related('department').get(
            id=student_id,
            courses__faculty=faculty
        )
        
        # Get courses the student is enrolled in that are taught by this faculty
        courses = student.courses.filter(faculty=faculty)
        
        # Calculate overall attendance
        attendance_records = AttendanceRecord.objects.filter(
            student=student,
            session__course__in=courses
        )
        total_sessions = attendance_records.count()
        present_sessions = attendance_records.filter(status=True).count()
        attendance_percentage = (present_sessions / total_sessions * 100) if total_sessions > 0 else 0
        
        # Get recent assignments
        recent_assignments = AssignmentSubmission.objects.filter(
            student=student,
            assignment__course__in=courses
        ).select_related(
            'assignment',
            'assignment__course'
        ).order_by('-submitted_at')[:5]
        
        context = {
            'student': student,
            'courses': courses,
            'attendance_percentage': round(attendance_percentage, 1),
            'total_sessions': total_sessions,
            'present_sessions': present_sessions,
            'recent_assignments': recent_assignments,
            'faculty': faculty
        }
        return render(request, 'faculty/student_details.html', context)
    except Student.DoesNotExist:
        messages.error(request, 'Student not found')
        return redirect('attendance:view_all_students')

@login_required
def student_attendance(request, student_id):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    faculty = request.user.faculty
    courses = Course.objects.filter(faculty=faculty)
    
    try:
        # Get the student directly using id
        student = Student.objects.get(id=student_id)
        
        # Verify that the student is enrolled in at least one of the faculty's courses
        if not courses.filter(students=student).exists():
            messages.error(request, 'Student not found in your courses')
            return redirect('attendance:faculty_dashboard')
        
        # Get attendance records for all courses
        attendance_records = AttendanceRecord.objects.filter(
            student=student,
            session__course__in=courses
        ).select_related('session__course', 'marked_by__user')
        
        # Use values to get unique dates first
        unique_dates = attendance_records.values_list('session__date', flat=True).distinct()
        
        # For each unique date, get the latest record
        latest_records = []
        for date in unique_dates:
            latest_record = attendance_records.filter(
                session__date=date
            ).order_by('-marked_at').first()
            if latest_record:
                latest_records.append(latest_record)
        
        # Sort the records by date
        latest_records.sort(key=lambda x: x.session.date, reverse=True)
        
        # Calculate attendance statistics per course
        course_stats = {}
        student_courses = courses.filter(students=student)
        for course in student_courses:
            course_records = [record for record in latest_records if record.session.course == course]
            total = len(course_records)
            present = sum(1 for record in course_records if record.status)
            course_stats[course.id] = {
                'total': total,
                'present': present,
                'absent': total - present,
                'percentage': round((present / total * 100) if total > 0 else 0, 1)
            }
        
        context = {
            'student': student,
            'attendance_records': latest_records,
            'course_stats': course_stats,
            'faculty': faculty,
            'courses': student_courses
        }
        return render(request, 'faculty/student_attendance.html', context)
        
    except Student.DoesNotExist:
        messages.error(request, 'Student not found')
        return redirect('attendance:faculty_dashboard')

@login_required
def student_grades(request, student_id):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    faculty = request.user.faculty
    courses = Course.objects.filter(faculty=faculty)
    student = get_object_or_404(Student, id=student_id, courses__in=courses)
    
    # Get all assignments and submissions for this student
    course_grades = []
    for course in courses.filter(students=student):
        assignments = Assignment.objects.filter(course=course)
        submissions = AssignmentSubmission.objects.filter(
            student=student,
            assignment__in=assignments
        ).select_related('assignment')
        
        total_marks = sum(sub.marks for sub in submissions if sub.marks is not None)
        total_assignments = assignments.count()
        
        course_grades.append({
            'course': course,
            'submissions': submissions,
            'total_marks': total_marks,
            'total_assignments': total_assignments,
            'average': round(total_marks / total_assignments if total_assignments > 0 else 0, 2)
        })
    
    context = {
        'student': student,
        'course_grades': course_grades
    }
    return render(request, 'faculty/student_grades.html', context)

@login_required
def delete_course(request, course_id):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    course = get_object_or_404(Course, id=course_id, faculty=request.user.faculty)
    
    if request.method == 'POST':
        course_name = course.name
        course.delete()
        messages.success(request, f'Course "{course_name}" has been deleted successfully')
        return redirect('attendance:manage_courses')
    
    return JsonResponse({'error': 'Invalid method'}, status=405)

@login_required
def view_recent_activities(request):
    if request.user.user_type != 'faculty':
        messages.error(request, 'Access denied')
        return redirect('attendance:login')
    
    faculty = request.user.faculty
    courses = Course.objects.filter(faculty=faculty)
    
    # Get recent attendance records with all necessary relations
    attendance_records = AttendanceRecord.objects.filter(
        session__course__in=courses
    ).select_related(
        'student__user',
        'session__course'
    ).order_by('-marked_at')[:20]
    
    # Get recent assignment submissions with all necessary relations
    assignment_submissions = AssignmentSubmission.objects.filter(
        assignment__course__in=courses
    ).select_related(
        'student__user',
        'assignment__course'
    ).order_by('-submitted_at')[:20]
    
    # Combine and sort activities
    recent_activities = sorted(
        chain(attendance_records, assignment_submissions),
        key=lambda x: x.marked_at if hasattr(x, 'marked_at') else x.submitted_at,
        reverse=True
    )[:30]  # Get top 30 most recent activities
    
    context = {
        'recent_activities': recent_activities,
        'faculty': faculty
    }
    return render(request, 'faculty/recent_activities.html', context)

@login_required
@csrf_protect
def verify_enrollment(request, course_id):
    """Verify if a student exists and is enrolled in a course."""
    logger.info(f"Verifying enrollment for course {course_id}")
    
    if request.user.user_type != 'faculty':
        logger.warning(f"Non-faculty user {request.user} attempted to verify enrollment")
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=405)
    
    try:
        course = get_object_or_404(Course, id=course_id, faculty=request.user.faculty)
        student_id = request.POST.get('student_id')
        
        if not student_id:
            logger.warning("Student ID not provided in request")
            return JsonResponse({'error': 'Student ID is required'}, status=400)
            
        student = Student.objects.filter(student_id=student_id).first()
        if not student:
            logger.info(f"Student with ID {student_id} not found")
            return JsonResponse({
                'exists': False,
                'message': 'Student not found'
            })
            
        is_enrolled = course.students.filter(id=student.id).exists()
        logger.info(f"Student {student_id} enrollment status for course {course_id}: {is_enrolled}")
        
        return JsonResponse({
            'exists': True,
            'enrolled': is_enrolled,
            'student': {
                'id': student.id,
                'name': student.user.get_full_name(),
                'student_id': student.student_id,
                'email': student.user.email
            }
        })
        
    except Course.DoesNotExist:
        logger.error(f"Course {course_id} not found")
        return JsonResponse({'error': 'Course not found'}, status=404)
    except Exception as e:
        logger.error(f"Error verifying enrollment: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)
